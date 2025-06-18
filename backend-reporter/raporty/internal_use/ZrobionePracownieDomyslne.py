import re
import os
import shutil
import base64
import datetime
import time
import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.drawing.image import Image
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, Color, PatternFill
from openpyxl.utils.units import pixels_to_points

from datasources.reporter import ReporterDatasource
from datasources.kakl import KaKlDatasource
from dialog import Dialog, Panel, HBox, VBox, TextInput, LabSelector, TabbedView, Tab, InfoText, DateInput, \
    Radio, ValidationError, Switch, BadanieSearch
from helpers.crystal_ball.marcel_cenniki_gotowkowe import wymuszony_cennik_domyslny
from helpers.validators import validate_date_range
from outlib.xlsx import ReportXlsx
from outlib.xlsx_standalone import RaportXlsxStandalone
from tasks import TaskGroup, Task
from helpers import prepare_for_json, get_centrum_connection, get_and_cache, first_or_none, divide_chunks, random_path, \
    copy_from_remote, ZIP, simple_password, list_from_space_separated

from raporty.Techniczne.PracownieDomyslne import raport_lista_badan as pd_raport_lista_badan, \
    raport_lab as pd_raport_lab
from raporty.Techniczne.CenyGotowkowe import raport_lab as cg_raport_lab, raport_snr as cg_raport_snr


# puszczać ok 9/10

"""
[19.12.22 12:00] Martyna Słowik
co tydzień z naszego działu wychodzi taki plik, który potocznie nazywamy pracownie domyślne

[19.12.22 12:01] Martyna Słowik
i mamy tutaj zestawienie trzech raportów: narzędzia techniczne 1. pracownie domyślne 2. pracownie domyślne + dołączyć czasy maksymalne na wykonanie 3. Ceny w cennikach gotówkowych

[19.12.22 12:01] Martyna Słowik
pytanie było takie, czy możemy zrobić taki raport, który będzie generować te trzy raporty w jednym?

[19.12.22 12:02] Martyna Słowik
Ania uważa, że to za bardzo czasochłonne żeby generowac trzy raporty i je łączyc

[19.12.22 12:02] Martyna Słowik
stąd taki pomysł jej 

[19.12.22 12:03] Martyna Słowik
w raporcie z czasami muszę scalać te dwie komórki 

[19.12.22 12:04] Martyna Słowik
dla każdego labu

[19.12.22 12:04] Martyna Słowik
w raporcie z cenami muszę ukrywać kolumny, które nie mają cennika domyślnego, ale to musi pozostać ręczną robotą bo czasami domyślny nie jest tym, który w rzeczywistości funkcjonuje 



Paulina Bereza weryfikacja:
- z zakładki Ceny wywalić te laby, w których nie ma cen
- lab orłowski - albo globalnie nieaktywny albo tu pomijać


"""

MENU_ENTRY = 'Pracownie domyślne'

LAUNCH_DIALOG = Dialog(title="Pracownie domyślne (3 raporty złożone do kupy)", panel=VBox(
    LabSelector(multiselect=True, field='laboratoria', title='Laboratorium'),
))


def start_report(params):
    params = LAUNCH_DIALOG.load_params(params)
    report = TaskGroup(__PLUGIN__, params)
    if len(params['laboratoria']) == 0:
        raise ValidationError("Nie wybrano żadnego laboratorium")
    lb_task = {
        'type': 'centrum',
        'priority': 1,
        'target': 'ZAWODZI',
        'params': {'badanie': None},
        'function': 'pd_raport_lista_badan',
    }
    report.create_task(lb_task)
    lb_task = {
        'type': 'snr',
        'priority': 1,
        'params': params,
        'function': 'cg_raport_snr',
    }
    report.create_task(lb_task)
    for lab in params['laboratoria']:
        lb_task = {
            'type': 'centrum',
            'priority': 1,
            'target': lab,
            'params': {'badanie': None},
            'function': 'pd_raport_lab',
        }
        report.create_task(lb_task)
        lb_task = {
            'type': 'centrum',
            'priority': 1,
            'target': lab,
            'params': {'badanie': None, 'ukryte': True,},
            'function': 'cg_raport_lab',
        }
        report.create_task(lb_task)
    report.save()
    return report


def zrob_raport(laby, badania, naglowek_badania, pracownie, aktywne, czasy, ceny, cenniki_domyslne) -> bytes:
    """
        badania = []
        naglowek_badania = [] # nazwa, grupa, grupa cennikowa
        pracownie = {} # lab -> badanie
        aktywne = {} # lab -> badanie
        czasy = {} # lab -> badanie
        ceny = {} # lab -> cennik -> badanie
        cenniki_domyslne = {} # lab -> cennik
    """

    xlsx = RaportXlsxStandalone(sheet="Pracownie domyślne")
    xlsx.set_columns(["Badanie", "Badanie nazwa", "Grupa", "Grupa cen."] + laby)
    xlsx.set_freeze(4, 1)
    for bad, nagl in zip(badania, naglowek_badania):
        row = [bad] + nagl
        for lab in laby:
            akt = aktywne[lab].get(bad, False)
            prac = pracownie[lab].get(bad, '')
            if akt:
                row.append(prac)
            else:
                row.append({'value': prac, 'background': '#cccccc'})
        xlsx.add_row(row)

    xlsx.add_sheet("Czas")
    header = ["Badanie", "Badanie nazwa", "Grupa", "Grupa cen."]
    for lab in laby:
        header.append({'title': lab, 'colspan': 2})
    xlsx.set_columns(header)
    xlsx.set_freeze(4, 1)
    for bad, nagl in zip(badania, naglowek_badania):
        row = [bad] + nagl
        for lab in laby:
            akt = aktywne[lab].get(bad, False)
            prac = pracownie[lab].get(bad, '')
            czas = czasy[lab].get(bad, '')
            if akt:
                row.append(prac)
                row.append(czas)
            else:
                row.append({'value': prac, 'background': '#cccccc'})
                row.append({'value': czas, 'background': '#cccccc'})
        xlsx.add_row(row)

    xlsx.add_sheet("Ceny")
    xlsx.set_columns(["Badanie", "Badanie nazwa", "Grupa", "Grupa cen."] + laby)
    xlsx.set_freeze(4, 1)
    for bad, nagl in zip(badania, naglowek_badania):
        row = [bad] + nagl
        for lab in laby:
            row.append(ceny.get(lab, {}).get(cenniki_domyslne[lab], {}).get(bad))
        xlsx.add_row(row)

    return xlsx.render_as_bytes()

    """
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Pracownie domyślne")
    ws.

    ws.freeze_panes = ws[0][3]

    ws = wb.create_sheet("Czas")

    ws = wb.create_sheet("Ceny")

    fn = random_path('reporter', 'xlsx')
    wb.save(fn)
    with open(fn, 'rb') as f:
        result = f.read()
    os.unlink(fn)
    return result
    """

def get_result(ident):
    task_group = TaskGroup.load(ident)
    if task_group is None:
        return None
    res = {
        'errors': [],
        'results': [],
    }
    lista_badan_result = None
    cenniki_snr_result = None
    pracownie_result = {}
    cenniki_domyslne = {}
    ceny_result = {}
    error_labs = set()
    start_params = None
    for job_id, params, status, result in task_group.get_tasks_results():
        if status == 'finished' and result is not None:
            # TODO: wrzucic błędy do błędów
            if params['function'] == 'pd_raport_lista_badan':
                # wiersze: symbol, grupa, grupa cennikowa (?), nazwa
                lista_badan_result = result
            if params['function'] == 'cg_raport_snr':
                # słownik lab -> słownik: symbol - (laboratorium symbol, del, zablokowany, uzywany, wycofany - info o cennikach)
                cenniki_snr_result = result
                start_params = params
            if params['function'] == 'pd_raport_lab':
                # metody_w_badaniach: lista słowników (badanie, maxy, pracownia),
                # badania_aktywne: lista słowników (badanie, bezr, ukryte, dorozliczen - integery)
                pracownie_result[params['target']] = result
            if params['function'] == 'cg_raport_lab':
                # ceny_w_badaniach: lista słowników (badanie, cena, cennik)
                ceny_result[params['target']] = result['ceny_w_badaniach']
        elif status == 'failed':
            error_labs.add(params['target'])
    res['progress'] = task_group.progress
    if task_group.progress == 1.0:
        laby = []
        badania = []
        naglowek_badania = [] # nazwa, grupa, grupa cennikowa
        pracownie = {} # lab -> badanie
        aktywne = {} # lab -> badanie
        czasy = {} # lab -> badanie
        ceny = {} # lab -> cennik -> badanie
        for poz in lista_badan_result:
            symbol = poz[0].strip()
            badania.append(symbol)
            naglowek_badania.append([poz[3], poz[1], poz[2]])
        for lab in start_params['params']['laboratoria']:
            if lab in error_labs:
                continue
            laby.append(lab)
            pracownie[lab] = {}
            aktywne[lab] = {}
            czasy[lab] = {}
            ceny[lab] = {}
            for poz in pracownie_result[lab]['metody_w_badaniach']:
                symbol = poz['badanie'].strip()
                pracownie[lab][symbol] = poz['pracownia']
                czasy[lab][symbol] = poz['maxy']
            for poz in pracownie_result[lab]['badania_aktywne']:
                symbol = poz['badanie'].strip()
                # print('poz', poz)
                aktywne[lab][symbol] = (poz['bezr'] or 0) + (poz['ukryte'] or 0) + (poz['dorozliczen'] or 0) == 0
            for poz in ceny_result[lab]:
                symbol = poz['badanie'].strip()
                cennik = poz['cennik'].strip()
                domyslny = 'domyślny' in cennik
                cennik = cennik.split(' ')[0]
                if cennik not in ceny[lab]:
                    ceny[lab][cennik] = {}
                    if domyslny:
                        cenniki_domyslne[lab] = cennik
                ceny[lab][cennik][symbol] = poz['cena']
            cenniki_domyslne[lab] = wymuszony_cennik_domyslny(lab, cenniki_domyslne.get(lab))
        res['results'].append({
            'type': 'download',
            'content': base64.b64encode(zrob_raport(laby, badania, naglowek_badania, pracownie, aktywne, czasy, ceny, cenniki_domyslne)).decode(),
            'content_type': "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            'filename': 'pracownie_domyslne_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d')
        })
    if len(error_labs) > 0:
        res['results'].append({
            'type': 'error',
            'text': 'Nie udało się pobrać danych z: %s' % ', '.join(list(error_labs))
        })
    return res
