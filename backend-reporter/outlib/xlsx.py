from pprint import pprint
import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.drawing.image import Image
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, Color, PatternFill
from openpyxl.utils.units import pixels_to_points
import webcolors
import os
import re
import datetime
from helpers import random_path
from .diagram import ReportDiagram

XLSX_ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')


class ReportXlsx:
    def __init__(self, data, **settings):

        self.data = data
        self.settings = {
            'flat_table': False,
            'flat_table_header': 'Tabela'
        }
        for k, v in settings.items():
            self.settings[k] = v
        self.global_header = [self.settings['flat_table_header']]
        self.global_header_titles = [
            self.cell_value_or_title(self.global_header[0])]
        self.row_count = 0
        # self.data: results, errors

    def get_color_value(self, value):
        if not value.startswith('#'):
            value = webcolors.name_to_hex(value)
        return value.replace('#', '')

    def collect_global_header(self, table_header):
        for cell in table_header:
            if self.cell_value_or_title(cell) not in self.global_header_titles:
                self.global_header.append(cell)
                self.global_header_titles.append(
                    self.cell_value_or_title(cell))

    @staticmethod
    def _cell_row_and_colspan(cell):
        if isinstance(cell, dict):
            rowspan = cell.get('rowspan', 1)
            colspan = cell.get('colspan', 1)
        else:
            rowspan = 1
            colspan = 1
        return rowspan, colspan

    def _max_header_columns(self, headers):
        results = []
        for r in headers:
            current_result = 0
            for c in r:
                rowspan, colspan = self._cell_row_and_colspan(c)
                current_result += colspan
            results.append(current_result)
        return max(results)

    def _header_template(self, data):
        matrix = []
        for r in range(len(data)):
            matrix_row = []
            for cell in range(self._max_header_columns(data)):
                matrix_row.append('e')
            matrix.append(matrix_row)
        return matrix

    @staticmethod
    def _clear_header_template(skipping):
        for row_index, row in enumerate(skipping):
            for cell_index, cell in enumerate(row):
                if cell == 'e' or cell == 't':
                    skipping[row_index][cell_index] = ''
        return skipping

    def _set_taken_cells_in_header_template(self, cell, starting_point, skipping):
        rowspan, colspan = self._cell_row_and_colspan(cell)
        for r in range(rowspan):
            for c in range(colspan):
                current_row = starting_point[0] + r
                current_col = starting_point[1] + c
                if (current_row, current_col) != starting_point:
                    skipping[current_row][current_col] = 't'
        return skipping

    def _set_values_in_header_template(self, data, skipping):

        for row_index, row in enumerate(data):
            for cell_index, cell in enumerate(row):
                for matrix_cell_index, matrix_cell in enumerate(
                        skipping[row_index]):
                    if matrix_cell == 'e':
                        skipping[row_index][matrix_cell_index] = 'v'
                        skipping = self._set_taken_cells_in_header_template(
                            cell, (row_index, matrix_cell_index), skipping)
                        break

        return skipping

    @staticmethod
    def _refactor_header_data(data):
        if not isinstance(data[0], list):
            return [data]
        return data

    def get_new_skipping(self, data):

        data = self._refactor_header_data(data)
        new_skipping = self._header_template(data)
        new_skipping = self._set_values_in_header_template(data, new_skipping)
        new_skipping = self._clear_header_template(new_skipping)

        return new_skipping

    def render_table_header(self, header, skipping=None):

        if header is None:
            return

        response_template = self.get_new_skipping(header)

        if not isinstance(header[0], list):
            header = [header]

        for header_row_indx, header_row in enumerate(header):
            for fld_indx, fld in enumerate(header_row):
                if isinstance(fld, str):
                    fld = {
                        'value': fld,
                        'fontstyle': 'b',
                    }
                for s_indx, s in enumerate(response_template[header_row_indx]):
                    if s == 'v':
                        c_val = self.formatted_cell(fld)
                        response_template[header_row_indx][s_indx] = c_val
                        break
        [self.add_row(row) for row in response_template]
        return

    def formatted_cell(self, cell):
        if isinstance(cell, dict):
            res = WriteOnlyCell(ws=self.ws, value=self._clean_value(
                cell.get('value', cell.get('title'))))
            font_attrs = {
                'name': 'Arial',
            }
            for ch in cell.get('fontstyle', ''):
                if ch == 'b':
                    font_attrs['b'] = True
                if ch == 'u':
                    font_attrs['u'] = 'single'
                if ch == 'i':
                    font_attrs['i'] = True
                if ch == 's':
                    font_attrs['strike'] = True
            if 'color' in cell:
                font_attrs['color'] = self.get_color_value(cell['color'])
            res.font = Font(**font_attrs)
            if 'background' in cell:
                res.fill = PatternFill(patternType='solid', fill_type='solid',
                                       fgColor=Color(self.get_color_value(
                                           cell['background'])))
            return res
        else:
            return self._clean_value(cell)

    def add_row(self, row):
        self.ws.append(row)
        self.row_count += 1

    def cell_value_or_title(self, cell):
        if isinstance(cell, dict):
            return cell.get('value', cell['title'])
        else:
            return cell

    def combine_table_header_with_global(self, header):
        header_titles = []
        global_titles = []
        result = []
        for tab, hdr in [
            [header_titles, header],
            [global_titles, self.global_header]
        ]:
            for cell in hdr:
                tab.append(self.cell_value_or_title(cell))
        for title in global_titles:
            res_title = None
            for i, h_title in enumerate(header_titles):
                if h_title == title:
                    res_title = i
            result.append(res_title)
        return result

    def render_table_title(self, title):
        cell = WriteOnlyCell(self.ws, value=self._clean_value(title))
        cell.font = Font(name='Arial', b=True, size=14)
        self.add_row([cell])

    def render_table_row(self, row):
        res = []
        for fld in row:
            res.append(self.formatted_cell(fld))
        self.add_row(res)

    def render_table_row_combined(self, title, row, combination):
        res = [title]
        for idx in combination[1:]:
            if idx is not None:
                res.append(self.formatted_cell(row[idx]))
            else:
                res.append('')
        self.add_row(res)

    def render_table(self, table):
        table_title = table.get('title') or ''
        if table_title != '' and not self.settings['flat_table']:
            self.render_table_title(table_title)
        if not self.settings['flat_table']:
            self.render_table_header(table.get('header'))

            for row in table['data']:
                self.render_table_row(row)
        else:
            combination = self.combine_table_header_with_global(
                table.get('header'))
            for row in table['data']:
                self.render_table_row_combined(table_title, row, combination)

    def render_vert_table(self, table):
        table_title = table.get('title') or ''
        if table_title != '':
            self.render_table_title(table_title)
        for row in table['data']:
            self.render_table_row([row['title'], row['value']])

    def render_info(self, info):
        cell = WriteOnlyCell(ws=self.ws, value=self._clean_value(info['text']))
        color = None
        b = False
        if info['type'] == 'info':
            color = self.get_color_value('darkblue')
        if info['type'] == 'warning':
            color = self.get_color_value('darkorange')
            b = True
        if info['type'] == 'error':
            color = self.get_color_value('darkred')
            b = True
        cell.font = Font(color=color, b=b)
        if self.settings['flat_table']:
            self.ws_par.append(['INFO', cell])
        else:
            self.add_row([cell])

    def render_diagram(self, diag):
        if 'title' in diag:
            self.render_table_title(diag['title'])
        next_row = self.row_count + 1
        diagram = ReportDiagram(diag)
        img_file = diagram.render_to_temp_file()
        img = Image(img_file)
        self.ws.row_dimensions[next_row].height = pixels_to_points(img.height)
        self.add_row([''])
        self.ws.add_image(img, anchor='A%d' % next_row)
        # os.unlink(img_file)
        # TODO: kasowanie pliku po wyrenderowaniu całości

    def _report_params_value(self, value):
        if isinstance(value, list):
            value = ', '.join([str(v) for v in value])
        return self._clean_value(value)

    def _render_report_params(self, params):
        if isinstance(params, dict):
            params = list(params.items())
        for row in params:
            cell = WriteOnlyCell(ws=self.ws_par,
                                 value=self._report_params_value(row[1]))
            cell.font = Font(b=True)
            self.ws_par.append([row[0], cell])

    def _clean_value(self, value):
        if isinstance(value, str):
            value = XLSX_ILLEGAL_CHARACTERS_RE.sub(' ', value)
        return value

    def render_to_file(self, fn):
        write_only = True
        if 'freeze_before' in self.settings:
            write_only = False
        self.wb = openpyxl.Workbook(write_only=write_only)
        if write_only:
            self.ws = self.wb.create_sheet('Dane')
        else:
            self.ws = self.wb.active
            self.ws.title = 'Dane'
        self.ws_par = self.wb.create_sheet('Parametry')
        if self.settings['flat_table']:
            for subres in self.data['results']:
                if subres['type'] == 'table':
                    if 'header' in subres:
                        self.collect_global_header(subres['header'])
            self.render_table_header(self.global_header)
        for subres in self.data['results']:
            if subres['type'] == 'table':
                self.render_table(subres)
            elif subres['type'] == 'vertTable':
                self.render_vert_table(subres)
            elif subres['type'] in ('info', 'warning', 'error'):
                self.render_info(subres)
            elif subres['type'] == 'diagram':
                self.render_diagram(subres)
            elif subres['type'] == 'download':
                self.render_info({'type': 'error', 'text': 'Nie można osadzić pliku do pobrania'})
            else:
                raise Exception('Unknown subres type', subres['type'])
        for suberr in self.data.get('errors', []):
            self.render_info({'type': 'error', 'text': suberr})
        if 'params' in self.data:
            self._render_report_params(self.data['params'])
        self.ws_par.append([])
        self.ws_par.append(
            [
                'Wygenerowano w aplikacji Alab Reporter %s' % datetime.datetime.now().strftime(
                    '%Y-%m-%d %H:%M')])
        if 'freeze_before' in self.settings:
            cell = self.ws[self.settings['freeze_before']]
            self.ws.freeze_panes = cell
        self.wb.save(fn)

    def render_as_bytes(self):
        fn = random_path('reporter', 'xlsx')
        self.render_to_file(fn)
        with open(fn, 'rb') as f:
            result = f.read()
        os.unlink(fn)
        return result
