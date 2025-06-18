import os

import webcolors
from openpyxl.styles import PatternFill, Color

from helpers import random_path
from .raport_standalone import RaportStandalone
from openpyxl import Workbook
from openpyxl.cell import Cell


"""
https://stackoverflow.com/questions/21875249/memory-error-using-openpyxl-and-large-data-excels
UPDATE: In version 2.4.0 use_iterators = True option is removed. In newer versions openpyxl.writer.write_only.WriteOnlyWorksheet is introduced for dumping large amounts of data.

from openpyxl import Workbook
wb = Workbook(write_only=True)
ws = wb.create_sheet()

# now we'll fill it with 100 rows x 200 columns
for irow in range(100):
    ws.append(['%d' % i for i in range(200)])

# save the file
wb.save('new_big_file.xlsx') 

"""

class RaportXlsxStandalone(RaportStandalone):
    def __init__(self, sheet=None, **kwargs):
        RaportStandalone.__init__(self, **kwargs)
        self.next_sheet_title = sheet
        self.wb = Workbook()
        self.ws = None
        self.freeze = None

    def to_excel(self, value):
        raw_value = value
        is_config_dict = False
        if isinstance(value, dict):
            raw_value = value['value']
            is_config_dict = True
        raw_value = RaportStandalone.to_excel(self, raw_value)
        cell = Cell(self.ws, value=raw_value)
        if is_config_dict and 'background' in value:
            cell.fill = PatternFill(patternType='solid', fill_type='solid',
                                   fgColor=Color(self._get_color_value(
                                       value['background'])))
        return cell

    def render_current_worksheet(self):
        self.prepare()
        self.ws = self.wb.active
        if self.next_sheet_title is not None:
            self.ws.title = self.next_sheet_title
        hdr_titles = []
        hdr_spans = []
        for col in self.col_titles:
            value = col
            colspan = 1
            if isinstance(col, dict):
                value = col.get('title', '')
                colspan = col.get('colspan', 1)
            hdr_titles.append(value)
            if colspan > 1:
                hdr_spans.append((len(hdr_titles), len(hdr_titles)+colspan-1))
                for _ in range(colspan-1):
                    hdr_titles.append('')
        self.ws.append(hdr_titles)
        for (start, end) in hdr_spans:
            self.ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        for row_num, row in enumerate(self.rows):
            self.ws.append(self.to_excel(value) for value in row)
        first_row = self.ws[1]
        for cell in first_row:
            cell.font = cell.font.copy(bold=True)
        if self.freeze is not None:
            cols_f, rows_f = self.freeze
            freeze_cell = self.ws[rows_f + 1][cols_f]
            self.ws.freeze_panes = freeze_cell

    def save_workbook_to_file(self, filename):
        self.wb.active = self.wb.worksheets[0]
        self.wb.save(filename)

    def render(self, filename=None):
        if filename is None:
            raise Exception('Brak nazwy pliku')
        self.render_current_worksheet()
        self.save_workbook_to_file(filename)

    def add_sheet(self, title):
        if self.cols is not None:
            self.render_current_worksheet()
            self.wb.active = self.wb.create_sheet(title)
            self.next_sheet_title = title
            self.reset()
            self.freeze = None
        else:
            self.wb.active.title = title

    def set_freeze(self, cols, rows):
        self.freeze = (cols, rows)

    def _get_color_value(self, value):
        if not value.startswith('#'):
            value = webcolors.name_to_hex(value)
        return value.replace('#', '')

    def render_as_bytes(self):
        fn = random_path('reporter', 'xlsx')
        self.render(fn)
        with open(fn, 'rb') as f:
            result = f.read()
        os.unlink(fn)
        return result
